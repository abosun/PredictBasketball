import torch
from torch import nn

import openpyxl
import pandas as pd
import numpy as np

EPOCH_NUM = 300

class NN2(nn.Module):
    """Neural Network model"""
    def __init__(self, input_size, output_size=1):
        super(NN2,self).__init__()
        self.fc1 = nn.Linear(input_size, 40)
        self.fc2 = nn.Linear(40, output_size)

    def forward(self,x):
        x = self.fc1(x)
        x = self.fc2(x)
        return x

def prepare_data():
    """read data from new.xlsx
       return: data(pd.dataframe)
               labels(pd.dataframe), 1: win, 0: lose
    """
    # prepare features
    data = pd.read_excel('./new.xlsx', sheet_name=1, usecols='I:M, O:AA', header=0)
    data = data.dropna()

    # prepare labels
    labels = pd.read_excel('./new.xlsx', sheet_name=1, usecols='E')
    labels = labels.dropna()
    labels = labels['单场胜负']
    for idx, label in labels.iteritems():
        if label == '胜':
            labels[idx] = 1
        elif label == '负':
            labels[idx] = 0
        else:
            assert False

    return data, labels


def prepare_data2():
    """read data from new.xlsx
       return: data(pd.dataframe)
    """
    # prepare features
    data = pd.read_excel('./new.xlsx', sheet_name=1, usecols='C, E, I:M, O:AA, CJ, CL, CP:CT, CV:DH', header=0)
    data = data.dropna()

    return data


def write_season_data():
    """read new.xlsx and extract season data, write to season_data.xlsx"""
    book = openpyxl.load_workbook("./new.xlsx")
    sheet = book.get_sheet_by_name('Sheet2')
    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row=row, column=2)
        print("cell is:", cell.value)  # could get 'dump', 'value', 'xf_index'
        print(cell.fill.start_color.rgb)
        

def prepare_win_lose_data(data):
    """input: data(pd.dataframe)
       return: team(pd.dataframe)
               labels(pd.dataframe)
    """
    team = data.iloc[:, [0, 20]]

    # prepare labels
    labels = data.iloc[:, 1]
    for idx, label in labels.iteritems():
        if label == '胜':
            labels[idx] = 1
        elif label == '负':
            labels[idx] = 0
        else:
            assert False
    return team, labels


def prepare_feature_data(data):
    """input: data(pd.dataframe)
       return: feature(pd.dataframe)
               labels(pd.dataframe)
    """
    team = data.iloc[:, [0, 20]]

    # prepare features
    features = data.iloc[:, 2:20]
    return team, features


def encodeteam(team):
    """encode team name to number"""
    encode = {
        '底特律活塞': 0,
        '密尔沃基雄鹿': 1,
        '金州勇士': 2,
        '洛杉矶快船': 3,
        '波士顿凯尔特人': 4, 
        '印第安纳步行者': 5, 
        '休斯顿火箭': 6,
        '犹他爵士': 7,
        '奥兰多魔术': 8,
        '多伦多猛龙': 9,
        '丹佛掘金': 10,
        '圣安东尼奥马刺': 11,
        '布鲁克林篮网': 12,
        '费城76人': 13, 
        '俄克拉荷马雷霆': 14,
        '波特兰开拓者': 15,
        '华盛顿奇才': 16, 
        '克里夫兰骑士': 17,
        '迈阿密热火': 18,
        '明尼苏达森林狼': 19,
        '新奥尔良鹈鹕': 20,
        '芝加哥公牛': 21,
        '亚特兰大老鹰': 22,
        '孟菲斯灰熊': 23,
        '夏洛特黄蜂': 24,
        '达拉斯小牛': 25,
        '夏洛特山猫': 26,
        '纽约尼克斯': 27,
        '洛杉矶湖人': 28,
        '菲尼克斯太阳': 29,
        '新泽西篮网': 30,
        '萨克拉门托国王': 31,
        '西雅图超音速': 32,
        '华盛顿子弹': 33,
        '对手': 34
    }
    team = team.replace(encode)
    return team


def split_data(data, labels):
    """split training set and testing set. 
       Write to xlsx file
    """
    idxs = np.random.randint(0, len(data), size=int(0.2*len(data)))
    train_data = data.drop(data.index[idxs])
    train_labels = labels.drop(labels.index[idxs])
    train_data.to_excel('./train_data.xlsx')
    train_labels.to_excel('./train_labels.xlsx')
    test_data = data.iloc[idxs]
    test_labels = labels.iloc[idxs]
    test_data.to_excel('./test_data.xlsx')
    test_labels.to_excel('./test_labels.xlsx')


def split_data2(data):
    """split training set and testing set. 
       Write to xlsx file
    """
    idxs = np.random.randint(0, len(data), size=int(0.2*len(data)))
    train_data = data.drop(data.index[idxs])
    train_data.to_excel('./train_data2.xlsx')
    test_data = data.iloc[idxs]
    test_data.to_excel('./test_data2.xlsx')


def load_train_test():
    """load data from train_data.xlsx and test_data.xlsx"""
    train_data   = pd.read_excel('./train_data.xlsx')
    train_data   = train_data.iloc[:, 1:]
    train_labels = pd.read_excel('./train_labels.xlsx', usecols='B')
    test_data    = pd.read_excel('./test_data.xlsx')
    test_data    = test_data.iloc[:, 1:]
    test_labels  = pd.read_excel('./test_labels.xlsx', usecols='B')
    return train_data, train_labels, test_data, test_labels


def train_win_lose(net, train_season_data, train_season_labels, train_offseason_data, train_offseason_labels):
    """training
       train_data: tensor
       train_labels: tensor
    """
    print('Start training...')
    criterion = nn.BCEWithLogitsLoss()
    # criterion = nn.CrossEntropyLoss()
    optimizer = torch.optim.Adam(net.parameters(), lr=1e-3)
    batchSize = 20
    import random
    for epoch in range(EPOCH_NUM):
        batchList = []
        for inputs, target in zip(train_season_data, train_season_labels):
            batchList.append([inputs,target])
        for inputs, target in zip(train_offseason_data, train_offseason_labels):
            batchList.append([inputs,target])
        random.shuffle(batchList)
        for i in range(len(batchList)//batchSize):
            batchData = batchList[batchSize*i: batchSize*(i+1)]
            inputs = torch.stack([x[0] for x in batchData],dim=0)
            target = torch.tensor([[x[1]] for x in batchData])
            # print(inputs.size())
            # target = torch.tensor([target])
            out = net(inputs)
            loss = criterion(out, target) * 0.8

            optimizer.zero_grad()
            loss.backward()
            optimizer.step()

        if (epoch + 1) % 100 == 0:
            print('Epoch: {}, Loss:{:.5f}'.format(epoch + 1, loss.item()))
    return net

def train(net, train_data, train_labels):
    """training
       train_data: tensor
       train_labels: tensor
    """
    print('Start training...')
    criterion = nn.BCEWithLogitsLoss()
    optimizer = torch.optim.Adam(net.parameters(), lr=1e-4)

    for epoch in range(300):
        for inputs, target in zip(train_data, train_labels):
            out = net(inputs)
            loss = criterion(out, target)

            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
        if (epoch + 1) % 100 == 0:
            print('Epoch: {}, Loss:{:.5f}'.format(epoch + 1, loss.item()))
    return net


def test(net, test_data, test_labels):
    """Test
       test_data: tensor
       test_labels: tensor
    """
    print('Testing...')
    running_corrects = 0
    positive = 0
    negative = 0
    TP = 0
    FP = 0
    TN = 0
    FN = 0
    for inputs, target in zip(test_data, test_labels):
        prob = net(inputs)
        if prob.item() > 0.5:
            pred = 1 
        else:
            pred = 0
        running_corrects += (pred == target.item())
        if target.item() == 1:
            positive += 1
            if pred == 1:
                TP += 1
            else:
                FP += 1
        elif target.item() == 0:
            negative += 1
            if pred == 1:
                FN += 1
            else:
                TN += 1
        else:
            assert False
    acc = running_corrects / len(test_labels)
    print('Accuracy:', acc)
    print('Positive sample:', positive)
    print('Negative sample:', negative)
    print('True Positive:', TP)
    print('True Negative:', TN)
    print('False Positive:', FP)
    print('False Negative:', FN)
    print('Precision:', TP/(TP+FP))
    print('Recall:', TP/(TP+FN))
    print('F1 Score:', 2*TP/(2*TP + FP + FN))
    return acc


def feature_input():
    """input feature, output win or lose"""
    train_data, train_labels, test_data, test_labels = load_train_test()
    train_data = train_data.to_numpy()
    train_labels = train_labels.to_numpy()
    train_data = torch.from_numpy(train_data).float()
    train_labels = torch.from_numpy(train_labels).float()

    test_data = test_data.to_numpy()
    test_labels = test_labels.to_numpy()
    test_data = torch.from_numpy(test_data).float()
    test_labels = torch.from_numpy(test_labels).float()

    net = NN2(input_size=18)
    net = train(net, train_data, train_labels)
    acc = test(net, test_data, test_labels)
    torch.save(net, './model-acc{:.2f}-v1.pkl'.format(acc))

def test_by_feature(test_data, test_labels):
    net = torch.load('./model-acc0.83-v1.pkl')
    running_corrects = 0
    positive = 0
    negative = 0
    TP = 0
    FP = 0
    TN = 0
    FN = 0
    probList = []
    for inputs, target in zip(test_data, test_labels):
        prob = net(inputs)
        probList.append(prob.item())
        if prob.item() > 0.5:
            pred = 1 
        else:
            pred = 0
        running_corrects += (pred == target.item())
        if target.item() == 1:
            positive += 1
            if pred == 1:
                TP += 1
            else:
                FP += 1
        elif target.item() == 0:
            negative += 1
            if pred == 1:
                FN += 1
            else:
                TN += 1
        else:
            assert False
    acc = running_corrects / len(test_labels)
    print('acc', acc)
    return acc, probList



def predict_by_feature(inputs):
    net = torch.load('./model-acc0.83-v1.pkl')
    prob = net(inputs)


if __name__ == '__main__':
    feature_input()